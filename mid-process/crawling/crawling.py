from openpyxl import load_workbook #엑셀 파일 읽기
import requests #html 받아오기
from bs4 import BeautifulSoup #html parsing 하기
import urllib.request #parsing한 링크로 이미지 다운로드
import urllib.parse #parsing한 링크에 있는 한글 변환
import os #image 담을 파일 생성

def uni_crawling(url, date):
    #requests로 html 받아오기(by http 응답)
    response = requests.get(url)

    #beautifulSoup로 html parsing 하기
    soup = BeautifulSoup(response.text, 'html.parser')

    #find는 html tag로 select는 css를 이용해 원하는 내용 찾음
    #find : 1개의 태그만 찾음.
    #find_all : 모든 태그를 찾음.
    #select_one : 1개의 태그만 찾음.
    #select : 모든 태그를 찾음.

    #parsing한 html에서 원하는 데이터 찾기
    
    #기사 제목
    titleContents = soup.select_one('.wrap_title_contents>h1')
    title = titleContents.text

    #이미지 주소
    temp = soup.select_one('.btn_download>a')
    href = temp.attrs['href']

    #이미지에 있는 한글을 unicode로 변환
    print(href)
    first = href.find("_")
    second = href.rfind("_")
    if first != second:
        print("error!!!")
        imgLink = href[:first+1] + urllib.parse.quote_plus(href[first+1:second]) + href[second:]
    else:
        imgLink = href

    #urlretrieve는 다운로드 함수
    urllib.request.urlretrieve(imgLink, "./images/"+ date + "_" + title + '.jpg')

#image 담을 file 생성
os.mkdir("images")

#excel 파일에서 url과 date 불러와 crawling 함수 실행
wb = load_workbook(filename = '(마스터) 칼럼 기고 목록(2020년~22.05.04) .xlsx')
sheet1 = wb['2020년']
sheet2 = wb['2021년']
sheet3 = wb['2022년']

sheets = [sheet1, sheet2, sheet3]

row = 0
for sheet in sheets:
    row = 3
    while(sheet.cell(row, 5).value != None):
        uni_crawling(sheet.cell(row=row, column=5).hyperlink.target, str((sheet.cell(row=row, column=1).value))[:10])
        row = row + 1