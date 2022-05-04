from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os

#출력파일 생성
output = Workbook()
sheetO = output.active

print("------------------------------------------------------------------------------")
print("크롤링 된 xlsx파일들의 데이터를 정해진 양식의 xlsx파일로 바꾸는 프로그램입니다")
print("(xls 형식은 지원하지 않습니다)\n")

print("xlsx파일들을 한 폴더에 넣은 후 폴더의 경로를 입력하면 한번에 처리합니다")
print("현재 폴더의 xlsx를 변환하고 싶으면 경로 입력없이 엔터를 누르면 됩니다\n")

print("xlsx파일들의 제목을 교수님의 성함으로 인식합니다")
print("크롤링을 진행할 때 미리 제목을 바꿔두세요")
print("------------------------------------------------------------------------------\n")

#파일 경로 입력받아 파일 리스트 생성
path = input("파일 경로 > ")
if (path[0] == "/" or path[0] == "\\"): path = path[1:]
file_list = os.listdir(os.path.join(".", path))

#output의 어느 row에서 입력을 시작할지
startrow = 1

#input에서 값을 읽어올때 사용하는 변수들
row = 2

#제대로 된 input인지 거르기 위한 list
classes = ["date", "title", "source", "contents", "link"]

for file in file_list:
    if (file[-5:] == ".xlsx") and (file[:2] != "~$"): #모든 파일중 .xlsx 형식만 통과, .xlsx 실행중 생기는 파일은 통과x
        
        #파일 읽어오기
        pathToFile = os.path.join(".",path,file)
        wb = load_workbook(filename = pathToFile)
        sheetI = wb.active

        #자료의 이름을 확인해 유효한 .xlsx인지 판단
        error = 0
        for col in range(5):
            if sheetI.cell(1, col+2).value != classes[col]: error=error+1
        if error > 0: continue

        #파일이름 작성
        sheetO.cell(row = startrow, column = 1, value= file )

        #date, 보도, 제목, 출처
        while (sheetI.cell(row, 2).value != None):
            #name
            sheetO.cell(startrow + row - 1, 1).value = file[:-5]
            sheetO.cell(startrow + row - 1, 1).alignment = Alignment(horizontal='center', vertical='center')
            #date
            sheetO.cell(startrow + row - 1, 4).value = int((sheetI.cell(row, 2).value).replace(".",""))
            sheetO.cell(startrow + row - 1, 4).alignment = Alignment(horizontal='center', vertical='center')
            #보도
            sheetO.cell(row = startrow + row - 1, column = 5, value= "보도" )
            sheetO.cell(startrow + row - 1, 5).value = "보도"
            sheetO.cell(startrow + row - 1, 5).alignment = Alignment(horizontal='center', vertical='center')
            #title
            sheetO.cell(startrow + row - 1, 6).value = sheetI.cell(row, 3).value
            sheetO.cell(startrow + row - 1, 6).hyperlink = sheetI.cell(row, 6).value
            #sheetO.cell(startrow + row - 1, 6).alignment = Alignment(horizontal='center', vertical='center')
            sheetO.cell(startrow + row - 1, 6).style = "Hyperlink"
            #출처
            sheetO.cell(startrow + row - 1, 7).value = sheetI.cell(row, 4).value
            sheetO.cell(startrow + row - 1, 7).alignment = Alignment(horizontal='left', vertical='bottom')

            row = row+1
            
        #다음 파일의 시작을 위한 세팅
        startrow += row
        row=2
        wb.close()

#출력
output.save(filename = "result.xlsx")