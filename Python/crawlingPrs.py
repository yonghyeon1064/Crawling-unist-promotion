from openpyxl import Workbook #엑셀 파일 쓰기
from openpyxl.styles import Alignment #엑셀 파일 셀 글자 위치 지정
import requests #html 받아오기
from bs4 import BeautifulSoup #html parsing 하기
import urllib.request #parsing한 링크로 이미지 다운로드
import urllib.parse #parsing한 링크에 있는 한글 변환
from datetime import datetime, timedelta #날짜 계산

def make_url(name, startDate, endDate, page):
    url = "https://search.naver.com/search.naver?where=news&sm=tab_pge&query=%EC%9C%A0%EB%8B%88%EC%8A%A4%ED%8A%B8%20%7C%20%EC%9A%B8%EC%82%B0%EA%B3%BC%EA%B8%B0%EC%9B%90%20%7C%20%EC%9A%B8%EC%82%B0%EA%B3%BC%ED%95%99%EA%B8%B0%EC%88%A0%EC%9B%90%20%7C%20unist%20%2B"
    url = url + urllib.parse.quote_plus(name) + "&sort=2&photo=0&field=0&pd=3"
    url = url + "&ds=" + startDate + "&de=" + endDate
    url = url + "&mynews=0&office_type=0&office_section_code=0&news_office_checked=&nso=so:r,p:from"
    url = url + startDate.replace(".", "") + "to" + endDate.replace(".", "") + ",a:all&start="
    page = str(int(page) - 1) + "1"
    if(page == "01"): page = "1"
    return url + page

def uni_crawling():

    #입력받기
    print("_____________________________________________________________________________\n")
    print("- 네이버 뉴스 중 UNIST 교수님 관련 기사들을 크롤링하는 프로그램입니다")
    print("- 양식에 맞게 입력하세요\n")
    print("- 이름 작성 완료 후 ..을 입력하면 result.xlsx 파일로 결과가 저장됩니다")
    print("_____________________________________________________________________________\n")

    sDate = input("시작날짜(ex: 2022.01.01): ")
    eDate = input("종료날짜(ex: 2022.01.02): ")
    maxPage = input("크롤링 최대 page: ")
    names = []
    name = "default"
    print("검색하고 싶은 교수님(ex: 홍길동)")
    while name != "..":
        name = input(">")
        if name=="..": break
        names.append(name)

    #output file 생성
    output = Workbook()
    sheetO = output.active

    #엑셀 row 포인터
    curRow = 1

    #crawling
    for name in names:
        for page in range(int(maxPage)):
            #url 생성
            url = make_url(name, sDate, eDate, page+1)

            #requests로 html 받아오기(by http 응답)
            response = requests.get(url)

            #beautifulSoup로 html parsing 하기
            soup = BeautifulSoup(response.text, 'html.parser')
            
            #html 페이지에 있는 모든 기사 묶음 list로 받아옴
            news_box = soup.select('.group_news>.list_news>.bx')

            #기사 묶음에서 원하는 정보들 추출 후 .xlsx 파일에 저장
            for i in news_box:
                #name
                sheetO.cell(curRow, 1).value = name
                sheetO.cell(curRow, 1).alignment = Alignment(horizontal='center', vertical='center')

                #title with hyperlink
                sheetO.cell(curRow, 6).value = i.select_one(".news_tit").attrs["title"]
                sheetO.cell(curRow, 6).hyperlink = i.select_one(".news_tit").attrs["href"]
                sheetO.cell(curRow, 6).style = "Hyperlink"

                #보도
                sheetO.cell(curRow, 5).value = "보도"
                sheetO.cell(curRow, 5).alignment = Alignment(horizontal='center', vertical='center')

                #온라인(default)
                sheetO.cell(curRow, 8).value = "온라인"

                info_list = i.select(".info_group>.info")
                for j in info_list:
                    if(j.find("span") != None): #출판사
                        sheetO.cell(curRow, 7).value = j.text
                        sheetO.cell(curRow, 7).alignment = Alignment(horizontal='left', vertical='bottom')
                    elif(j.find("i") != None): #출판면(온라인 갱신)
                        sheetO.cell(curRow, 8).value = j.text
                        sheetO.cell(curRow, 8).alignment = Alignment(horizontal='left', vertical='center')                        
                    elif(j.text != "네이버뉴스"): #date
                        date = (j.text).replace(".","")
                        if date.isdigit():
                            sheetO.cell(curRow, 4).value = int(date)
                        else:
                            timeDiff = -1 * int(date[0])
                            articleTime = datetime.now() + timedelta(days=timeDiff)
                            sheetO.cell(curRow, 4).value = int(articleTime.strftime("%Y%m%d"))
                        sheetO.cell(curRow, 4).alignment = Alignment(horizontal='center', vertical='center')
                
                curRow = curRow + 1
        curRow = curRow + 1

    output.save(filename = "result.xlsx")

uni_crawling()