from openpyxl import load_workbook
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb.active
print(sheet_ranges.cell(18, 4).value)